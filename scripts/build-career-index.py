"""
build-career-index.py — turns the owner's Grand Audio Career Index workbook
into the two bundled data files the Audio Career Finder ships with.

    py scripts/build-career-index.py "C:\\path\\to\\Pro_Audio_Training_Academy_Grand_Audio_Career_Index-2.xlsx"

Writes:
    src/data/careerIndex.json      — every career title, trimmed + enum-coded
    src/data/careerFamilies.json   — per-family metadata derived from the sheet
                                      (settings, title count, curriculum links)

The 42 families' names, dimensions, examples and descriptions are hand-authored
in src/features/careerfinder/families.ts; this script only supplies what the
workbook knows. Family rows are matched by exact name, so a renamed family in
either place fails loudly here rather than silently in the app.

Curriculum links are resolved against the LIVE v3 curriculum (REST, anon key
from .env) so every gs number in the output is a real active topic. Topic names
that do not resolve (the workbook sometimes names a SUBJECT as a topic) are
dropped — the family still links to its subject.
"""
import io, json, os, re, sys, urllib.request
from collections import OrderedDict

import openpyxl

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SRC = sys.argv[1] if len(sys.argv) > 1 else os.path.join(os.path.expanduser('~'), 'Downloads', 'Pro_Audio_Training_Academy_Grand_Audio_Career_Index-2.xlsx')
OUT_CAREERS = os.path.join(ROOT, 'src', 'data', 'careerIndex.json')
OUT_FAMILIES = os.path.join(ROOT, 'src', 'data', 'careerFamilies.json')
V3_ID = 'a7c1f2e0-9b34-4d55-8e21-0c4f6a9b1d72'

def norm(s):
    s = (s or '').lower().replace('&', 'and')
    return re.sub(r'[\s\u2014\u2013\-/:]+', ' ', s).strip()

def slug(s):
    s = s.lower().replace('&', 'and')
    s = re.sub(r'[^a-z0-9]+', '-', s).strip('-')
    return s

def env():
    out = {}
    for f in ('.env', '.env.local'):
        p = os.path.join(ROOT, f)
        if os.path.exists(p):
            for line in io.open(p, encoding='utf-8'):
                m = re.match(r'\s*([A-Z_]+)\s*=\s*"?([^"\n]+)"?', line)
                if m: out[m.group(1)] = m.group(2).strip()
    return out

def fetch_v3():
    e = env()
    url = e.get('EXPO_PUBLIC_SUPABASE_URL'); key = e.get('EXPO_PUBLIC_SUPABASE_ANON_KEY')
    if not url or not key:
        sys.exit('EXPO_PUBLIC_SUPABASE_URL / ANON_KEY missing from .env')
    q = f"{url}/rest/v1/achievements?curriculum_version_id=eq.{V3_ID}&is_active=eq.true&select=global_sequence,name,field,subject&order=global_sequence"
    req = urllib.request.Request(q, headers={'apikey': key, 'Authorization': 'Bearer ' + key})
    return json.load(urllib.request.urlopen(req))

# ── enumerations (index = code stored in JSON) ────────────────────────────
TIER = ['Direct audio occupation', 'Audio-dependent allied profession', 'Context-specific paid role']
RELATIONSHIP = ['Core audio/acoustics', 'Closely related / supporting', 'Audio-dependent allied work', 'Acoustics-based profession', 'Adjacent but audio-specific']
ORIENTATION = ['Technical / engineering', 'Hands-on technical / trade', 'Business / editorial / operations', 'Creative / performance', 'Science / research', 'Education / coaching', 'Clinical / therapeutic', 'Legal / policy / standards']
TITLE_CLASS = ['Established industry role', 'Technical/creative specialty', 'Operational/creative role', 'Leadership/management', 'Research specialty', 'Education/communication', 'Business/legal/adjacent role', 'Consulting/freelance service', 'Skilled trade/craft', 'Regulated or credentialed profession', 'Entry/support role']
STATUS = ['Established occupation or industry role', 'Established specialty or employer title', 'Emerging or fast-evolving title', 'Regulated/credentialed title']
WORK_MODEL = ['Staff roles common; consulting/contract options vary', 'Staff, freelance, or project-based depending on sector', 'Often freelance/contract; staff roles also exist', 'Project, venue, tour, or contract work common']

def code(table, value, what):
    v = (value or '').strip()
    if v not in table:
        sys.exit(f'unknown {what}: {v!r}')
    return table.index(v)

def main():
    wb = openpyxl.load_workbook(SRC, read_only=True, data_only=True)
    v3 = fetch_v3()
    by_name = {norm(t['name']): t for t in v3}
    subjects = {(t['field'], t['subject']) for t in v3}

    # sources sheet: S01 → organization name, so the app never shows a bare code
    src_rows = list(wb['Sources & Method'].iter_rows(values_only=True))
    source_name = {}
    for r in src_rows:
        if r and isinstance(r[0], str) and re.match(r'^S\d\d$', r[0]) and r[1]:
            source_name[r[0]] = re.sub(r'(\s+-\s+|\s*[—–]\s*).*$', '', str(r[1])).strip()  # 'BLS - Audiologists' -> 'BLS'; hyphenated names survive

    # families sheet
    rows = list(wb['Career Families'].iter_rows(values_only=True))
    fam_rows = [r for r in rows[5:] if r and r[0]]
    families = OrderedDict()
    unresolved = []
    for r in fam_rows:
        name, count, field, subject, topics, settings, sources = r[0], r[1], r[2], r[3], r[4], r[5], r[6]
        # The workbook's "Primary app subject" is the v3 FIELD and its
        # "Primary app field" is the v3 SUBJECT (the app's tree is Field → Subject).
        if (field, subject) not in subjects:
            sys.exit(f'{name}: {field} › {subject} is not a live v3 field/subject')
        gs = []
        for t in [x.strip() for x in str(topics).split(';') if x.strip()]:
            hit = by_name.get(norm(t))
            if hit: gs.append(hit['global_sequence'])
            else: unresolved.append((name, t))
        families[name] = {
            'id': slug(name),
            'name': name,
            'field': field,
            'subject': subject,
            'topicGs': gs,
            'settings': [s.strip() for s in str(settings).split(';') if s.strip()],
            'sources': list(OrderedDict.fromkeys(source_name.get(s.strip(), s.strip()) for s in str(sources).split(';') if s.strip())),
            'count': 0,
        }
    fam_ids = list(families.keys())

    # master list
    rows = list(wb['Career Master List'].iter_rows(values_only=True))
    hdr_i = next(i for i, r in enumerate(rows) if r and sum(1 for c in r if c) >= 8)
    hdr = [str(h) for h in rows[hdr_i]]
    col = {h: i for i, h in enumerate(hdr)}
    careers = []
    for r in rows[hdr_i + 1:]:
        if not r or not r[0]: continue
        g = lambda h: r[col[h]]
        fam = g('Career family')
        if fam not in families: sys.exit(f'{g("Career ID")}: unknown family {fam!r}')
        families[fam]['count'] += 1
        alt = g('Alternate / search titles')
        careers.append({
            'id': g('Career ID'),
            'f': fam_ids.index(fam),
            't': str(g('Career title')).strip(),
            **({'alt': [a.strip() for a in str(alt).split(';') if a.strip()]} if alt else {}),
            'cls': code(TITLE_CLASS, g('Title class'), 'title class'),
            'st': code(STATUS, g('Title status'), 'title status'),
            'rel': code(RELATIONSHIP, g('Relationship to audio'), 'relationship'),
            'tier': code(TIER, g('Inclusion tier'), 'tier'),
            'ori': code(ORIENTATION, g('Primary orientation'), 'orientation'),
            'wm': code(WORK_MODEL, g('Typical work model'), 'work model'),
            'prep': str(g('Typical preparation pathway')).strip(),
            **({'reg': 1} if g('Licensure / caution note') else {}),
        })

    # ── corrections on top of the workbook (scripts/career-index-overrides.json) ──
    ov = json.load(io.open(os.path.join(ROOT, 'scripts', 'career-index-overrides.json'), encoding='utf-8'))
    by_title = {}
    for c in careers: by_title.setdefault(c['t'], []).append(c)
    def each(title):
        rows = by_title.get(title)
        if not rows: sys.exit(f'override names a title not in the workbook: {title!r}')
        return rows
    for title, fam in ov['moveToFamily'].items():
        for c in each(title):
            families[fam_ids[c['f']]]['count'] -= 1
            c['f'] = fam_ids.index(fam)
            families[fam]['count'] += 1
    for title in ov['regulated']:
        for c in each(title): c['reg'] = 1
    pe_fams = {fam_ids.index(f) for f in ov['professionalEngineerFamilies']}
    for c in careers:
        if c['f'] in pe_fams and re.search(r'(Engineer|Consultant)$', c['t']): c['pe'] = 1
    for title, prep in ov['preparation'].items():
        for c in each(title): c['prep'] = prep
    DEGREE_DEFAULT = "Bachelor's degree common; portfolio/certification may supplement"
    HANDS_ON = 'Hands-on training, certificate/associate degree, apprenticeship, or experience'
    VARIES = "Varies: portfolio/experience, certificate, associate, or bachelor's pathway"
    craft = {fam_ids.index(f) for f in ov['craftFamilies']}
    for c in careers:
        if c['f'] in craft and c['prep'] == DEGREE_DEFAULT:
            c['prep'] = HANDS_ON if re.search(r'(Technician|Operator|Assistant|Stagehand|Tech|Crew|Hand|Roadie|Runner)\b', c['t']) else VARIES
    for title, ori in ov['orientation'].items():
        if title.startswith('_'): continue
        for c in each(title): c['ori'] = code(ORIENTATION, ori, 'orientation')
    clinical = code(ORIENTATION, 'Clinical / therapeutic', 'orientation')
    for c in careers:
        if 'Audiologist' in c['t']: c['ori'] = clinical
    for title, cls in ov['titleClass'].items():
        for c in each(title): c['cls'] = code(TITLE_CLASS, cls, 'title class')
    demote = {RELATIONSHIP.index(r) for r in ov['demoteCoreWhenRelationshipIn']}
    demoted = 0
    for c in careers:
        if c['tier'] == 0 and c['rel'] in demote:
            c['tier'] = 1; demoted += 1
    print(f'overrides: {len(ov["moveToFamily"])} moved, {len(ov["regulated"])} regulated, {sum(1 for c in careers if c.get("pe"))} PE-flagged, {demoted} demoted core->specialized')

    # preparation pathway is 14 distinct strings — table it too
    preps = sorted({c['prep'] for c in careers})
    for c in careers: c['prep'] = preps.index(c['prep'])

    out = {
        'version': 'grand-audio-career-index-v2-2026-09-04',
        'families': fam_ids,
        'enums': {'tier': TIER, 'relationship': RELATIONSHIP, 'orientation': ORIENTATION, 'titleClass': TITLE_CLASS, 'status': STATUS, 'workModel': WORK_MODEL, 'preparation': preps},
        'careers': careers,
    }
    io.open(OUT_CAREERS, 'w', encoding='utf-8', newline='\n').write(json.dumps(out, ensure_ascii=False, separators=(',', ':')) + '\n')
    io.open(OUT_FAMILIES, 'w', encoding='utf-8', newline='\n').write(json.dumps(list(families.values()), ensure_ascii=False, indent=1) + '\n')
    print(f'careers: {len(careers)}  families: {len(families)}  -> {os.path.relpath(OUT_CAREERS, ROOT)} ({os.path.getsize(OUT_CAREERS)//1024} KB), {os.path.relpath(OUT_FAMILIES, ROOT)}')
    print(f'topic links resolved: {sum(len(f["topicGs"]) for f in families.values())}; unresolved names dropped: {len(unresolved)}')
    for fam, t in unresolved: print(f'   {fam[:40]:40s} x {t}')

if __name__ == '__main__':
    main()
